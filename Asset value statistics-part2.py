#%%
import os
import openpyxl
import datetime
import xlrd
import chardet
import sys

# Function to detect encoding of a file
def detect_file_encoding(file_path):
    with open(file_path, 'rb') as file:
        raw_data = file.read()
        encoding = chardet.detect(raw_data)
        return encoding['encoding']

def remove_specific_lines(filename, encoding):
    unwanted_starts = {"东恒量化1号", "喜悦量化私募", "喜悦量化1号"}

    with open(filename, 'r', encoding=encoding) as file:
        lines = file.readlines()

    # Filter out lines that are empty, contain only whitespace, or start with specific strings
    filtered_lines = [line for line in lines if line.strip() and not any(line.startswith(s) for s in unwanted_starts)]

    # Write the filtered lines back to the file
    with open(filename, 'w', encoding=encoding) as file:
        file.writelines(filtered_lines)

def find_value_in_column_xlsx(worksheet, column_number, search_text):
    for row in worksheet.iter_rows(min_col=column_number, max_col=column_number):
        for cell in row:
            if cell.value == search_text:
                return cell.row
    return None

def find_value_in_column_xlsx_length(worksheet, column_number, search_text, required_length=None):
    for row in worksheet.iter_rows(min_col=column_number, max_col=column_number):
        for cell in row:
            cell_value = cell.value
            if required_length is not None:
                # Check if the cell's value matches the search text and has the required length
                if search_text in str(cell_value) and len(str(cell_value)) == required_length:
                    return cell.row
            else:
                if cell_value == search_text:
                    return cell.row
    return None


def find_value_in_row_xlsx(worksheet, row_number, search_text):
    for cell in worksheet[row_number]:
        if cell.value == search_text:
            return cell.column
    return None

def find_value_in_row_xlsx_length(worksheet, row_number, search_text, required_length=None):
    for cell in worksheet[row_number]:
        cell_value = cell.value
        if required_length is not None:
            # Check if the cell's value matches the search text and has the required length
            if search_text in str(cell_value) and len(str(cell_value)) == required_length:
                return cell.column
        else:
            if cell_value == search_text:
                return cell.column
    return None


def find_value_in_column_xls(sheet, col_idx, search_text):
    for row_idx in range(sheet.nrows):
        cell_value = sheet.cell_value(row_idx, col_idx)  # Check the first column
        if search_text in cell_value:
            return row_idx
    return None

def find_value_in_row_xls(sheet, row_idx, search_text):
    for col_idx in range(sheet.ncols):
        cell_value = sheet.cell_value(row_idx, col_idx)
        if search_text in str(cell_value):  # Convert cell value to string
            return col_idx
    return None

def find_value_in_column_xls_length(sheet, col_idx, search_text, required_length):
    for row_idx in range(sheet.nrows):
        cell_value = sheet.cell_value(row_idx, col_idx)
        if isinstance(cell_value, str) and search_text in cell_value and len(cell_value) == required_length:
            return row_idx
    return None

#get Current working directory
current_date = datetime.datetime.now().strftime('%Y-%m-%d')

path = os.getcwd()+ "\\" + current_date + "\\"

print('path:',path)
# Get the current date in the format YYYY-MM-DD



# Path to the folder containing Excel files with the date as the folder name
folder = os.path.exists(path)
if not folder:
    print("目标当日日期净值文件夹不存在，请按回车键退出：")
    input();sys.exit()

# Define the name of the existing text file based on today's date
existing_txt_file_name = f"净值汇总{current_date}.txt"

existing_txt_file_path = os.path.join(path, existing_txt_file_name)

file_names = sorted(os.listdir(path))


existing_txt_file_encoding = detect_file_encoding(existing_txt_file_path)

if existing_txt_file_encoding is None:
    existing_txt_file_encoding = 'utf-8'

remove_specific_lines(existing_txt_file_path, existing_txt_file_encoding)


# Check if the folder exists
if not os.path.exists(path):
    print(f"Folder for {current_date} does not exist.")
else:
    with open(existing_txt_file_path, 'a', encoding=existing_txt_file_encoding) as output_file:
    # Iterate through each file in the folder
        for filename in file_names:
            if filename.endswith('.xlsx') and "晨跃1号" in filename:
                # Construct the full file path
                file_path = os.path.join(path, filename)

                # Load the workbook and select the active worksheet
                workbook = openpyxl.load_workbook(file_path)
                worksheet = workbook.active

                # Store the output values in a list
                output_values = ["晨跃1号"]

                # Get date in 4th row, 1st column and extract the part after "日期:"
                date_cell = worksheet.cell(row=4, column=1)
                date_cell_value = date_cell.value if date_cell else ''
                date_str = date_cell_value.split("日期：")[-1].strip() if date_cell_value and "日期：" in date_cell_value else str(date_cell_value)
                output_values.append(date_str)

                # Get value to the right of "今日单位净值"
                today_net_value_row = find_value_in_column_xlsx(worksheet, 1, "今日单位净值")
                if today_net_value_row:
                    today_net_value = worksheet.cell(row=today_net_value_row, column=2).value


                if isinstance(today_net_value, float):       
                    rounded_today_net_value = round(today_net_value, 4)
                    value_str = str(rounded_today_net_value)
                else:
                    value_str = str(rounded_today_net_value)
                
                output_values.append(value_str)


                # Get value to the right of "累计单位净值"
                cumulative_net_value_row = find_value_in_column_xlsx(worksheet, 1, "累计单位净值")
                if cumulative_net_value_row:
                    cumulative_net_value = worksheet.cell(row=cumulative_net_value_row, column=2).value

                if isinstance(cumulative_net_value , float):       
                    cumulative_net_value  = round(cumulative_net_value , 4)
                    value_str = str(cumulative_net_value)
                else:
                    value_str = str(cumulative_net_value)
                output_values.append(value_str)


                # Get value of "资产净值" under "市值" column
                asset_net_value_row = find_value_in_column_xlsx(worksheet, 1, "资产净值")
                market_value_column = find_value_in_row_xlsx(worksheet, 5, "市值")
                if asset_net_value_row and market_value_column:
                    asset_net_value = worksheet.cell(row=asset_net_value_row, column=market_value_column).value
                    output_values.append(str(asset_net_value))

                # Print the collected values
                output_file.write(' '.join(output_values) + '\n')
                if filename.endswith('.xls'):
                    workbook.release_resources()

            elif filename.endswith('.xls') and "估值表_安值岭盛量化1号" in filename:
                # Construct the full file path
                file_path = os.path.join(path, filename)

                # Open the workbook
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)

                # Store the output values in a list
                output_values = ["岭盛1号"]

                # Get date in 4th row, 1st column and extract the part after "日期:"
                date_cell_value = sheet.cell_value(2, 1)  # 4th row (index 3), 2nd column (index 1)
                date_str = date_cell_value.split("日期：")[-1].strip() if "日期：" in date_cell_value else date_cell_value
                output_values.append(date_str)

                # Get value to the right of "基金单位净值："
                today_net_value_row = find_value_in_column_xls_length(sheet, 1, "基金单位净值", 7)
                if today_net_value_row is not None:
                    today_net_value = sheet.cell_value(today_net_value_row, 2)  # Next column
                    output_values.append(str(today_net_value))

                # Get value to the right of "累计单位净值："
                cumulative_net_value_row = find_value_in_column_xls_length(sheet, 1, "累计单位净值",7)
                if cumulative_net_value_row is not None:
                    cumulative_net_value = sheet.cell_value(cumulative_net_value_row, 2)  # Next column
                    output_values.append(str(cumulative_net_value))

                # Get value of "基金资产净值：" under "市值" column
                asset_net_value_row = find_value_in_column_xls_length(sheet, 1, "基金资产净值",7)
                market_value_column = find_value_in_row_xls(sheet, 3, "市值")  # 5th row (index 4)
                if asset_net_value_row is not None and market_value_column is not None:
                    asset_net_value = sheet.cell_value(asset_net_value_row, market_value_column)
                    output_values.append(str(asset_net_value))

                # Print the collected values
                output_file.write(' '.join(output_values) + '\n')

                if filename.endswith('.xls'):
                    workbook.release_resources()


            elif filename.endswith('.xls') and "国联大方向" in filename:
                file_path = os.path.join(path, filename)

                # Open the workbook
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)

                # Store the output values in a list
                output_values = ["国联大方向"]

                # Get date in 4th row, 1st column and extract the part after "日期:"
                date_cell_value = sheet.cell_value(2, 1)  # 3rd row (index 2), 2nd column (index 1)
                date_str = date_cell_value.split("日期：")[-1].strip() if "日期：" in date_cell_value else date_cell_value
                output_values.append(date_str)

                # Get value to the right of "基金单位净值："
                today_net_value_row = find_value_in_column_xls_length(sheet, 1, "基金单位净值", 7)
                if today_net_value_row is not None:
                    today_net_value = sheet.cell_value(today_net_value_row, 2)  # Next column
                    output_values.append(str(today_net_value))

                # Get value to the right of "累计单位净值："
                cumulative_net_value_row = find_value_in_column_xls_length(sheet, 1, "累计单位净值",7)
                if cumulative_net_value_row is not None:
                    cumulative_net_value = sheet.cell_value(cumulative_net_value_row, 2)  # Next column
                    output_values.append(str(cumulative_net_value))

                # Get value of "基金资产净值：" under "市值" column
                asset_net_value_row = find_value_in_column_xls_length(sheet, 1, "基金资产净值",7)
                market_value_column = find_value_in_row_xls(sheet, 3, "市值")  # 5th row (index 4)
                if asset_net_value_row is not None and market_value_column is not None:
                    asset_net_value = sheet.cell_value(asset_net_value_row, market_value_column)
                    output_values.append(str(asset_net_value))

                # Print the collected values
                output_file.write(' '.join(output_values) + '\n')
                if filename.endswith('.xls'):
                    workbook.release_resources()

            #处理中盛1B
            elif filename.endswith('.xls') and "中盛量化1号私募" in filename:
                file_path = os.path.join(path, filename)

                # Open the workbook
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)

                # Store the output values in a list
                output_values = ["中盛1号B类份额"]

                # Get date in 4th row, 1st column and extract the part after "日期:"
                date_cell_value = sheet.cell_value(3, 0)  # 4th row (index 3), first column (index 0)
                date_str = date_cell_value.split("日期：")[-1].strip() if "日期：" in date_cell_value else date_cell_value
                output_values.append(date_str)

                # Get value to the right of "基金单位净值："
                today_net_value_row = find_value_in_column_xls(sheet, 0, "今日单位净值") + 2 #(+2才是B类的位置)
                if today_net_value_row is not None:
                    today_net_value = sheet.cell_value(today_net_value_row, 1)  # Next column
                    output_values.append(str(today_net_value))

                # Get value to the right of "累计单位净值："
                cumulative_net_value_row = find_value_in_column_xls(sheet, 0, "累计单位净值")+2
                if cumulative_net_value_row is not None:
                    cumulative_net_value = sheet.cell_value(cumulative_net_value_row, 1)  # Next column
                    output_values.append(str(cumulative_net_value))

                # Get value of "基金资产净值：" under "市值" column
                asset_net_value_row = find_value_in_column_xls(sheet, 0, "资产净值")
                market_value_column = find_value_in_row_xls(sheet, 4, "市值") + 1 # 5th row (index 4)（本币）
                if asset_net_value_row is not None and market_value_column is not None:
                    asset_net_value = sheet.cell_value(asset_net_value_row, market_value_column)
                    output_values.append(str(asset_net_value))

                # Print the collected values
                output_file.write(' '.join(output_values) + '\n')

                if filename.endswith('.xls'):
                    workbook.release_resources()       

            #喜悦量化1号
            elif filename.endswith('.xls') and "喜悦量化1号" in filename:
                file_path = os.path.join(path, filename)

                # Open the workbook
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)

                # Store the output values in a list
                output_values = ["喜悦量化1号"]

                # Get date in 4th row, 1st column and extract the part after "日期:"
                date_cell_value = sheet.cell_value(2, 7)  # 3rd row (index 2), 8th column (index 7)
                date_str = date_cell_value.split("日期：")[-1].strip() if "日期：" in date_cell_value else date_cell_value
                output_values.append(date_str)

                # Get value to the right of "基金单位净值："
                today_net_value_row = find_value_in_column_xls(sheet, 0, "基金单位净值") 
                if today_net_value_row is not None:
                    today_net_value = sheet.cell_value(today_net_value_row, 1)  # Next column
                    output_values.append(str(today_net_value))

                # Get value to the right of "累计单位净值："
                cumulative_net_value_row = find_value_in_column_xls(sheet, 0, "累计单位净值")
                if cumulative_net_value_row is not None:
                    cumulative_net_value = sheet.cell_value(cumulative_net_value_row, 1)  # Next column
                    output_values.append(str(cumulative_net_value))

                # Get value of "基金资产净值：" under "市值" column
                asset_net_value_row = find_value_in_column_xls(sheet, 0, "基金资产净值")
                market_value_column = find_value_in_row_xls(sheet, 3, "市值")  # 4th row (index 3)（本币）
                if asset_net_value_row is not None and market_value_column is not None:
                    asset_net_value = sheet.cell_value(asset_net_value_row, market_value_column)
                    output_values.append(str(asset_net_value))

                # Print the collected values
                output_file.write(' '.join(output_values) + '\n')

                if filename.endswith('.xls'):
                    workbook.release_resources()                 

            #喜悦量化1号
            elif filename.endswith('.xls') and "喜悦量化私募" in filename:
                file_path = os.path.join(path, filename)

                # Open the workbook
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)

                # Store the output values in a list
                output_values = ["喜悦量化私募"]

                # Get date in 4th row, 1st column and extract the part after "日期:"
                date_cell_value = sheet.cell_value(2, 7)  # 3rd row (index 2), 8th column (index 7)
                date_str = date_cell_value.split("日期：")[-1].strip() if "日期：" in date_cell_value else date_cell_value
                output_values.append(date_str)

                # Get value to the right of "基金单位净值："
                today_net_value_row = find_value_in_column_xls_length(sheet, 0, "基金单位净值",7) 
                if today_net_value_row is not None:
                    today_net_value = sheet.cell_value(today_net_value_row, 1)  # Next column
                    output_values.append(str(today_net_value))

                # Get value to the right of "累计单位净值："
                cumulative_net_value_row = find_value_in_column_xls_length(sheet, 0, "累计单位净值",7)
                if cumulative_net_value_row is not None:
                    cumulative_net_value = sheet.cell_value(cumulative_net_value_row, 1)  # Next column
                    output_values.append(str(cumulative_net_value))

                # Get value of "基金资产净值：" under "市值" column
                asset_net_value_row = find_value_in_column_xls(sheet, 0, "基金资产净值")
                market_value_column = find_value_in_row_xls(sheet, 3, "市值")  # 4th row (index 3)（本币）
                if asset_net_value_row is not None and market_value_column is not None:
                    asset_net_value = sheet.cell_value(asset_net_value_row, market_value_column)
                    output_values.append(str(asset_net_value))

                # Print the collected values
                output_file.write(' '.join(output_values) + '\n')

                if filename.endswith('.xls'):
                    workbook.release_resources()       

            #东恒1好
            elif filename.endswith('.xls') and "东恒量化1号" in filename:
                file_path = os.path.join(path, filename)

                # Open the workbook
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)

                # Store the output values in a list
                output_values = ["东恒量化1号"]

                # Get date in 4th row, 1st column and extract the part after "日期:"
                date_cell_value = sheet.cell_value(2, 0)  # 3rd row (index 2), first column (index 0)
                date_str = date_cell_value.split("日期：")[-1].strip() if "日期：" in date_cell_value else date_cell_value
                output_values.append(date_str)

                # Get value to the right of "基金单位净值："
                today_net_value_row = find_value_in_column_xls_length(sheet, 0, "基金单位净值",7) 
                if today_net_value_row is not None:
                    today_net_value = sheet.cell_value(today_net_value_row, 1)  # Next column
                    output_values.append(str(today_net_value))

                # Get value to the right of "累计单位净值："
                cumulative_net_value_row = find_value_in_column_xls_length(sheet, 0, "累计单位净值",7)
                if cumulative_net_value_row is not None:
                    cumulative_net_value = sheet.cell_value(cumulative_net_value_row, 1)  
                    output_values.append(str(cumulative_net_value))

                # Get value of "基金资产净值：" under "市值" column
                asset_net_value_row = find_value_in_column_xls_length(sheet, 0, "基金资产净值",7)
                market_value_column = find_value_in_row_xls(sheet, 3, "市值")
                if asset_net_value_row is not None and market_value_column is not None:
                    asset_net_value = sheet.cell_value(asset_net_value_row, market_value_column)
                    output_values.append(str(asset_net_value))

                # Print the collected values
                output_file.write(' '.join(output_values) + '\n')

                if filename.endswith('.xls'):
                    workbook.release_resources()      

            #东恒1号
            elif filename.endswith('.xls') and "青峰量化1号私募" in filename:
                file_path = os.path.join(path, filename)

                # Open the workbook
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)

                # Store the output values in a list
                output_values = ["青峰1号"]

                # Get date in 4th row, 1st column and extract the part after "日期:"
                date_cell_value = sheet.cell_value(2, 0)  # 3rd row (index 2), first column (index 0)
                date_str = date_cell_value.split("日期：")[-1].strip() if "日期：" in date_cell_value else date_cell_value

                # Check if the date_str is in the expected format and then format it
                if len(date_str) == 8 and date_str.isdigit():
                    formatted_date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
                else:
                    formatted_date_str = date_str  # Use the original date_str if it's not in the expected format

                output_values.append(formatted_date_str)


                # Get value to the right of "基金单位净值："
                today_net_value_row = find_value_in_column_xls_length(sheet, 0, "基金单位净值",7) 
                if today_net_value_row is not None:
                    today_net_value = sheet.cell_value(today_net_value_row, 1)  # Next column
                    output_values.append(str(today_net_value))

                # Get value to the right of "累计单位净值："
                cumulative_net_value_row = find_value_in_column_xls_length(sheet, 0, "累计单位净值",7)
                if cumulative_net_value_row is not None:
                    cumulative_net_value = sheet.cell_value(cumulative_net_value_row, 1)  
                    output_values.append(str(cumulative_net_value))

                # Get value of "基金资产净值：" under "市值" column
                asset_net_value_row = find_value_in_column_xls_length(sheet, 0, "基金资产净值",7)
                market_value_column = find_value_in_row_xls(sheet, 3, "市值")
                if asset_net_value_row is not None and market_value_column is not None:
                    asset_net_value = sheet.cell_value(asset_net_value_row, market_value_column)
                    output_values.append(str(asset_net_value))

                # Print the collected values
                output_file.write(' '.join(output_values) + '\n')

                if filename.endswith('.xls'):
                    workbook.release_resources()   


            elif filename.endswith('.xls') and "青峰量化1号A期" in filename:
                file_path = os.path.join(path, filename)

                # Open the workbook
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)

                # Store the output values in a list
                output_values = ["青峰1号A期"]

                # Get date in 4th row, 1st column and extract the part after "日期:"
                date_cell_value = sheet.cell_value(2, 0)  # 3rd row (index 2), first column (index 0)
                date_str = date_cell_value.split("日期：")[-1].strip() if "日期：" in date_cell_value else date_cell_value

                # Check if the date_str is in the expected format and then format it
                if len(date_str) == 8 and date_str.isdigit():
                    formatted_date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
                else:
                    formatted_date_str = date_str  # Use the original date_str if it's not in the expected format

                output_values.append(formatted_date_str)

                # Get value to the right of "基金单位净值："
                today_net_value_row = find_value_in_column_xls(sheet, 0, "基金单位净值") 
                if today_net_value_row is not None:
                    today_net_value = sheet.cell_value(today_net_value_row, 1)  # Next column
                    output_values.append(str(today_net_value))

                # Get value to the right of "累计单位净值："
                cumulative_net_value_row = find_value_in_column_xls(sheet, 0, "累计单位净值")
                if cumulative_net_value_row is not None:
                    cumulative_net_value = sheet.cell_value(cumulative_net_value_row, 1)  # Next column
                    output_values.append(str(cumulative_net_value))

                # Get value of "基金资产净值：" under "市值" column
                asset_net_value_row = find_value_in_column_xls(sheet, 0, "基金资产净值")
                market_value_column = find_value_in_row_xls(sheet, 3, "市值")  # 4th row (index 3)（本币）
                if asset_net_value_row is not None and market_value_column is not None:
                    asset_net_value = sheet.cell_value(asset_net_value_row, market_value_column)
                    output_values.append(str(asset_net_value))

                # Print the collected values
                output_file.write(' '.join(output_values) + '\n')

                if filename.endswith('.xls'):
                    workbook.release_resources()   




            elif filename.endswith('.xls') and "盈玖量化1号" in filename:
                file_path = os.path.join(path, filename)

                # Open the workbook
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)

                # Store the output values in a list
                output_values = ["盈玖1号"]

                # Get date in 4th row, 1st column and extract the part after "日期:"
                date_cell_value = sheet.cell_value(2, 0)  # 3rd row (index 2), first column (index 0)
                date_str = date_cell_value.split("日期：")[-1].strip() if "日期：" in date_cell_value else date_cell_value

                # Check if the date_str is in the expected format and then format it
                if len(date_str) == 8 and date_str.isdigit():
                    formatted_date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
                else:
                    formatted_date_str = date_str  # Use the original date_str if it's not in the expected format

                output_values.append(formatted_date_str)

                # Get value to the right of "基金单位净值："
                today_net_value_row = find_value_in_column_xls_length(sheet, 0, "基金单位净值",7) 
                if today_net_value_row is not None:
                    today_net_value = sheet.cell_value(today_net_value_row, 1)  # Next column
                    output_values.append(str(today_net_value))

                # Get value to the right of "累计单位净值："
                cumulative_net_value_row = find_value_in_column_xls_length(sheet, 0, "累计单位净值",7)
                if cumulative_net_value_row is not None:
                    cumulative_net_value = sheet.cell_value(cumulative_net_value_row, 1)  # Next column
                    output_values.append(str(cumulative_net_value))

                # Get value of "基金资产净值：" under "市值" column
                asset_net_value_row = find_value_in_column_xls_length(sheet, 0, "基金资产净值",7)
                market_value_column = find_value_in_row_xls(sheet, 3, "市值")  # 4th row (index 3)（本币）
                if asset_net_value_row is not None and market_value_column is not None:
                    asset_net_value = sheet.cell_value(asset_net_value_row, market_value_column)
                    output_values.append(str(asset_net_value))

                # Print the collected values
                output_file.write(' '.join(output_values) + '\n')

                if filename.endswith('.xls'):
                    workbook.release_resources()  


            elif filename.endswith('.xlsx') and "青宁1号" in filename:
                # Construct the full file path
                file_path = os.path.join(path, filename)

                # Load the workbook and select the active worksheet
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                worksheet = workbook.active

                # Store the output values in a list
                output_values = ["青宁1号"]


                date_column = find_value_in_row_xlsx(worksheet, 3, "估值日期：") +  1

                date_cell = worksheet.cell(row=3, column=date_column)
                date_cell_value = date_cell.value if date_cell.value is not None else ''
                date_str = str(date_cell_value)

                output_values.append(date_str)


                # Get value to the right of "今日单位净值"
                today_net_value_row = find_value_in_column_xlsx_length(worksheet, 1, "今日单位净值",7)
                if today_net_value_row:
                    today_net_value = worksheet.cell(row=today_net_value_row, column=2).value
                    
                if isinstance(today_net_value, float):       
                    rounded_today_net_value = round(today_net_value, 4)
                    value_str = str(rounded_today_net_value)
                else:
                    value_str = str(rounded_today_net_value)
                
                output_values.append(value_str)

                ###
                #cell_value_test = worksheet.cell(row=152, column=2).value
                #print(cell_value_test)

                # Get value to the right of "累计单位净值"
                cumulative_net_value_row = find_value_in_column_xlsx_length(worksheet, 1, "累计单位净值",7)
                if cumulative_net_value_row:
                    cumulative_net_value = worksheet.cell(row=cumulative_net_value_row, column=2).value

                if isinstance(cumulative_net_value , float):       
                    cumulative_net_value  = round(cumulative_net_value , 4)
                    value_str = str(cumulative_net_value)
                else:
                    value_str = str(cumulative_net_value)

                output_values.append(value_str)

                # Get value of "资产净值" under "市值" column
                asset_net_value_row = find_value_in_column_xlsx_length(worksheet, 1, "资产资产净值",7)
                market_value_column = find_value_in_row_xlsx(worksheet, 4, "市 值")
                if asset_net_value_row and market_value_column:
                    asset_net_value = worksheet.cell(row=asset_net_value_row, column=market_value_column).value
                    output_values.append(str(asset_net_value))

                # Print the collected values
                output_file.write(' '.join(output_values) + '\n')
                if filename.endswith('.xls'):
                    workbook.release_resources()



# Function to remove empty lines from a file
def remove_empty_lines(filename, encoding):
    with open(filename, 'r', encoding=encoding) as file:
        lines = file.readlines()

    # Filter out lines that are empty or contain only whitespace
    non_empty_lines = [line for line in lines if line.strip()]

    # Write the non-empty lines back to the file
    with open(filename, 'w', encoding=encoding) as file:
        file.writelines(non_empty_lines)

# Detect the encoding of the existing file
existing_txt_file_encoding = detect_file_encoding(existing_txt_file_path)

# Fallback to 'utf-8' if no encoding is detected
if existing_txt_file_encoding is None:
    existing_txt_file_encoding = 'utf-8'

# Call the function to remove empty lines
remove_empty_lines(existing_txt_file_path, existing_txt_file_encoding)

# %%
